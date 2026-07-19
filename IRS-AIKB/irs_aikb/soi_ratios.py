"""IRS SOI aggregate corporation-ratio context.

SOI publication tables contain aggregate estimates, not return-level peer
distributions.  Results from this module are cohort context and never an audit
probability, percentile, or proof of an error.
"""
from __future__ import annotations

from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


ITEMS = {
    "return_count": "Number of returns",
    "total_assets": "Total assets",
    "cash": "Cash",
    "accounts_receivable": "Trade notes and accounts receivable",
    "inventory": "Inventories",
    "other_current_assets": "Other current assets",
    "accounts_payable": "Accounts payable",
    "short_term_debt": "Mortgages, notes, bonds payable in less than 1 year",
    "other_current_liabilities": "Other current liabilities",
    "total_receipts": "Total receipts",
    "business_receipts": "Business receipts",
    "total_deductions": "Total deductions",
    "cost_of_goods_sold": "Cost of goods sold",
    "officer_compensation": "Compensation of officers",
    "salaries_wages": "Salaries and wages",
    "rent": "Rents paid",
    "depreciation": "Depreciation",
    "advertising": "Advertising",
    "net_income": "Net income (less deficit) from a trade or business",
}


def _text(value: Any) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def _number(value: Any) -> float | None:
    if isinstance(value, bool) or value in (None, "", "**"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.()-]", "", str(value))
    if not cleaned:
        return None
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    try:
        number = float(cleaned.strip("()"))
        return -number if negative else number
    except ValueError:
        return None


def _ratio(numerator: float | None, denominator: float | None) -> float | None:
    return None if numerator is None or denominator in (None, 0) else numerator / denominator


def load_1120s_industry_ratios(workbook: Path, industry: str) -> dict:
    """Load an industry column from Publication 16 Table 6.1."""
    sheet = load_workbook(workbook, read_only=True, data_only=True).active
    requested = _text(industry).casefold()
    industry_column = None
    current_sector = ""
    selected_label = ""
    for column in range(2, sheet.max_column + 1):
        sector = _text(sheet.cell(5, column).value)
        detail = _text(sheet.cell(6, column).value)
        if sector:
            current_sector = sector
        label = detail if detail and detail.casefold() != "total" else current_sector
        if requested == label.casefold() or requested == current_sector.casefold():
            industry_column = column
            selected_label = label
            if requested == label.casefold():
                break
    if industry_column is None:
        raise ValueError(f"industry not found in SOI table: {industry}")
    row_by_label = {_text(sheet.cell(row, 1).value): row for row in range(1, sheet.max_row + 1)}
    values = {key: _number(sheet.cell(row_by_label[label], industry_column).value)
              for key, label in ITEMS.items() if label in row_by_label}
    receipts = values.get("business_receipts")
    total_receipts = values.get("total_receipts")
    current_assets = sum(value or 0 for value in (values.get("cash"),
        values.get("accounts_receivable"), values.get("inventory"),
        values.get("other_current_assets")))
    current_liabilities = sum(value or 0 for value in (values.get("accounts_payable"),
        values.get("short_term_debt"), values.get("other_current_liabilities")))
    ratios = {
        "gross_margin": _ratio((receipts - values["cost_of_goods_sold"])
            if receipts is not None and values.get("cost_of_goods_sold") is not None else None, receipts),
        "net_margin": _ratio(values.get("net_income"), total_receipts),
        "officer_compensation_to_receipts": _ratio(values.get("officer_compensation"), total_receipts),
        "wages_to_receipts": _ratio(values.get("salaries_wages"), total_receipts),
        "rent_to_receipts": _ratio(values.get("rent"), total_receipts),
        "advertising_to_receipts": _ratio(values.get("advertising"), total_receipts),
        "depreciation_to_receipts": _ratio(values.get("depreciation"), total_receipts),
        "inventory_to_assets": _ratio(values.get("inventory"), values.get("total_assets")),
        "receivables_to_receipts": _ratio(values.get("accounts_receivable"), receipts),
        "current_ratio": _ratio(current_assets, current_liabilities),
    }
    year_match = re.match(r"(\d{2})co", workbook.name.lower())
    tax_year = 2000 + int(year_match.group(1)) if year_match else None
    return {"benchmark_type": "IRS_SOI_AGGREGATE_COHORT",
            "tax_year": tax_year, "form_family": "1120-S",
            "industry": selected_label, "return_count_estimate": values.get("return_count"),
            "ratios": ratios, "source": {"table": "Publication 16 Table 6.1",
            "file": workbook.name,
            "official_url": f"https://www.irs.gov/pub/irs-soi/{workbook.name}"},
            "limitations": ["Aggregate estimates are not return-level peer observations.",
                "The ratios are not IRS audit-selection thresholds or percentiles.",
                "SOI sampling, disclosure, classification, and comparability limitations apply."]}


def compare_to_soi(taxpayer_ratios: dict[str, float | None], cohort: dict) -> dict:
    comparisons = []
    for name, taxpayer_value in taxpayer_ratios.items():
        soi_value = cohort["ratios"].get(name)
        if taxpayer_value is None or soi_value is None:
            comparisons.append({"ratio": name, "status": "not_comparable"})
            continue
        difference = taxpayer_value - soi_value
        comparisons.append({"ratio": name, "taxpayer": taxpayer_value,
            "soi_aggregate": soi_value, "difference": difference,
            "relative_difference": None if soi_value == 0 else difference / abs(soi_value),
            "status": "review_context", "risk_effect": "requires_professional_interpretation"})
    return {"comparison_type": "aggregate_context", "cohort": cohort,
            "comparisons": comparisons,
            "warning": "Difference from an SOI aggregate is a review signal, not proof of error or audit selection."}
